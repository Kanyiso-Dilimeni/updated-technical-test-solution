from django.shortcuts import render

from .models import Customer
from .forms import CustomerForm
import pandas as pd
import openpyxl as xl


def index(request):
    months = []
    income = []
    expenses = []

    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            file_name = str(request.FILES['customer_excel_file'].name)  # Gets the file name of the uploaded file
            new_file_name = file_name.replace(' ', '_')  # Returns name where the whitespaces are replaced with
            # underscores.

            workbook = xl.load_workbook(f'media/files/{new_file_name}', data_only=True)
            sheet = workbook['Sheet1']

            for row in range(2, sheet.max_row + 1):
                month = sheet.cell(row, 1)
                month_income = sheet.cell(row, 2)
                expense = sheet.cell(row, 3)

                months.append(month.value)
                income.append(month_income.value)
                expenses.append(expense.value)

            print(months)
            return render(request, 'finances/index.html', context={'form': form,
                                                                   'months': months,
                                                                   'income': income,
                                                                   'expenses': expenses})
    else:
        form = CustomerForm()
        print(months)
    context = {
        'form': form,
        'months': months,
        'income': income,
        'expenses': expenses
    }
    return render(request, 'finances/index.html', context)
